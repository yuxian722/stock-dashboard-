"""Minimal web UI wrapping the bingomap core.

Flow matches the original ask: open the page -> fill in header info ->
generate the blank skeleton -> click/drag-select wafer coordinates ->
generate the final .strate file. All business logic (blank generation,
quantity validation, DIE_INFO assignment) stays in the bingomap package;
this module is just HTTP plumbing + a thin static frontend.

Wafer bin data (the green/magenta grid) still has no automatic source —
see bingomap/README.md — so this UI accepts it as pasted "x,y,bin" text
for now. Swapping that input for a real data source later doesn't touch
anything below /api/generate.
"""
from __future__ import annotations

import base64
import binascii
import csv
import io
import json
import os
import re
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request
from openpyxl import Workbook, load_workbook

from bingomap.assignment import DieCountMismatch, DiePick, assign_dies, assign_layers
from bingomap.blank_generator import blank_from_positions, generate_blank
from bingomap.crack_recovery import (
    MissingNotchError,
    build_session,
    crack_csv_rows,
    crack_output_coord,
    wafer_scatter,
)
from bingomap.frm_reader import FrmFormatError, frm_file_path, frm_to_wafer_bin_map, parse_frm
from bingomap.mispick_analysis import (
    DECISION_ANOMALY,
    DECISION_FORCE_DELETE,
    DECISION_OK,
    DECISION_REVIEW,
    InvalidGeometryError,
    UnknownMachineTypeError,
    UnsupportedNotchError,
    analyze_substrate,
    make_offset,
    output_coord,
    parse_bin_set,
)
from bingomap.secs_log import decode_secs_log, extract_strate_files, extract_wafer_maps
from bingomap.secs_params import (
    ChecklistRow,
    SecsParam,
    SecsParamsFormatError,
    compare_checklist,
    extract_pp_param_snapshots,
)
from bingomap.strate import StrateFile, StrateFormatError

app = Flask(__name__)

# Matches WaferCoordinate.exe.config's own FRM_PATH default. Only reachable
# when this Flask process itself runs on a machine with the F: network
# drive mapped — see bingomap/CLAUDE.md. Override with the BINGOMAP_FRM_PATH
# env var, or the frm_path field in a /api/frm request body, for other
# deployments.
DEFAULT_FRM_PATH = os.environ.get("BINGOMAP_FRM_PATH", r"F:\SMAP\FRM\\")

# ⑥ SECS格式化參數頁的基準參數清單 — 2026/08/19使用者要求「把這頁改成
# 這199格式化參數固定在裡面」：不要每次都要重新上傳log才看得到，而是
# 把已經從真實log解析出來的199組參數(RECIPE@AEU132X2C001A-2070,
# TID=58151)當作固定基準值直接內建在頁面裡。之後使用者會提供Excel
# 參數清單，用這份基準值跟Excel清單做check list式比對(哪些已經在機台裡、
# 哪些是Excel裡有、機台還沒有、後續要慢慢加進去的349組擴充目標)。
SECS_PARAMS_BASELINE_PATH = Path(__file__).resolve().parent.parent / "bingomap" / "data" / "secs_params_baseline.json"


def _load_secs_params_baseline() -> dict:
    with open(SECS_PARAMS_BASELINE_PATH, encoding="utf-8") as f:
        return json.load(f)


def _blank_from_header(data: dict):
    return generate_blank(
        assy_lot=data["assy_lot"],
        mapping_lot=data["mapping_lot"],
        eqpid=data["eqpid"],
        oper=data["oper"],
        substrate_id=data["substrate_id"],
        substrate_row=int(data["substrate_row"]),
        substrate_column=int(data["substrate_column"]),
        substrate_block=int(data["substrate_block"]),
        notch=data.get("notch", ""),
        ref=data.get("ref", ""),
        t2_point=data.get("t2_point", "NA"),
        t2_flat=data.get("t2_flat", "NA"),
        out_mgz_slot_no=data.get("out_mgz_slot_no", ""),
        convention=data.get("convention", "EPOXY"),
        machine_type=data.get("machine_type", "DB"),
    )


@app.get("/")
def index():
    return render_template("index.html", active_page="supplement")


@app.post("/api/blank")
def api_blank():
    data = request.get_json(force=True)
    try:
        blank = _blank_from_header(data)
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify(
        {
            "positions": [d.sub_pos for d in blank.die_info],
            "total_qty": blank.total_bond_die_qty,
        }
    )


def _die_info_to_picks(die_list: list) -> list[dict]:
    picks = []
    for d in die_list:
        x_str, _, y_str = d.wafer_xy.partition(":")
        try:
            x, y = int(x_str), int(y_str)
        except ValueError:
            continue
        picks.append({"x": x, "y": y, "bin": d.bin})
    return picks


def _split_into_layer_picks(template: StrateFile) -> list[list[dict]]:
    """Group a parsed template's dies back into assign_layers()'s
    layer_picks shape: one pick-list per layer, ascending by f9, with the
    DIE_INFO layer (the highest f9 — see assign_layers()) last.

    [DIE_INFO_OTHER_LAYER_*] can hold several layers mixed together,
    distinguished only by each row's own f9 (confirmed against a real
    8-layer sample — see bingomap/tests/test_strate_eight_layer_real_sample.py),
    so this groups by f9 rather than assuming one section = one layer."""
    if not template.die_info:
        return []
    if not template.other_layer_die_info:
        return [_die_info_to_picks(template.die_info)]

    by_f9: dict[str, list] = {}
    for d in template.other_layer_die_info:
        by_f9.setdefault(d.f9, []).append(d)
    ordered_f9 = sorted(by_f9, key=lambda f9: int(f9))
    layers = [_die_info_to_picks(by_f9[f9]) for f9 in ordered_f9]
    layers.append(_die_info_to_picks(template.die_info))
    return layers


@app.post("/api/parse_strate")
def api_parse_strate():
    """複製既有.strate為範本：parse an existing real file and hand back
    everything the frontend needs to prefill the form and picks — header
    fields, the file's own substrate position order (verbatim, so
    regenerating never has to re-guess convention/machine_type), and the
    wafer picks already made, split by layer if the file has a stacked
    OTHER_LAYER section."""
    data = request.get_json(force=True)
    text = data.get("text", "")
    if not text.strip():
        return jsonify({"error": "請提供.strate檔案內容"}), 400

    try:
        template = StrateFile.parse(text)
    except StrateFormatError as exc:
        return jsonify({"error": f"檔案格式解析失敗：{exc}"}), 422

    return jsonify(
        {
            "assy_lot": template.assy_lot,
            "mapping_lot": template.mapping_lot,
            "eqpid": template.eqpid,
            "oper": template.oper,
            "substrate_id": template.substrate_id,
            "substrate_row": template.substrate_row,
            "substrate_column": template.substrate_column,
            "substrate_block": template.substrate_block,
            "notch": template.notch,
            "ref": template.ref,
            "wafer_ring": template.die_info[0].wafer_ring if template.die_info else "",
            "positions": [d.sub_pos for d in template.die_info],
            "total_qty": len(template.die_info),
            "picks": _die_info_to_picks(template.die_info),
            "num_layers": 1 + len({d.f9 for d in template.other_layer_die_info}),
            "layer_picks": _split_into_layer_picks(template),
        }
    )


@app.post("/api/frm")
def api_frm():
    """Auto-load the wafer bin map straight from the FRM file — the
    replacement for pasting "x,y,bin" text, once this server itself runs
    somewhere with F:\\SMAP\\FRM\\ mapped."""
    data = request.get_json(force=True)
    lot_no = (data.get("lot_no") or "").strip()
    barcode_id = (data.get("barcode_id") or "").strip()
    frm_root = (data.get("frm_path") or DEFAULT_FRM_PATH).strip()
    if not lot_no or not barcode_id:
        return jsonify({"error": "lot_no 和 barcode_id 都必填"}), 400

    try:
        path = frm_file_path(frm_root, lot_no, barcode_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    # frm_file_path() always joins with backslashes to match
    # WaferCoordinate.exe's own path construction (and to display a
    # Windows-familiar path in error messages) — real Windows deployments
    # handle that natively, but a non-Windows filesystem (e.g. this
    # session's Linux test suite) needs it translated to os.sep to
    # actually resolve.
    fs_path = path.replace("\\", os.sep)

    if not os.path.exists(fs_path):
        return (
            jsonify(
                {
                    "error": f"找不到檔案：{path}\n"
                    "請確認 LotNo/Barcode ID 是否正確，或這台電腦是否連得到F槽網路磁碟機"
                }
            ),
            404,
        )

    try:
        with open(fs_path, "rb") as f:
            frm = parse_frm(f.read())
    except FrmFormatError as exc:
        return jsonify({"error": f"檔案格式解析失敗：{exc}"}), 422
    except OSError as exc:
        return jsonify({"error": f"讀取檔案失敗：{exc}"}), 500

    return jsonify(
        {
            "columns": frm.col,
            "rows": frm.row,
            "lot_no": frm.lot_no,
            "wafer_id": frm.wafer_id,
            "wafer_type": frm.wafer_type,
            "reference_point_x": frm.reference_point_x,
            "reference_point_y": frm.reference_point_y,
            "cells": [{"x": x, "y": y, "bin": str(bin_kind)} for (x, y), bin_kind in frm.die_map.items()],
        }
    )


def _build_picks(selections: list[dict], sub_positions: list[str], wafer_ring: str) -> list[DiePick]:
    # assign_dies()/assign_two_layers() check len(picks) against
    # expected_qty before ever looking at an individual pick's sub_pos, so
    # when counts don't match we only need the right *length* to get a
    # correct DieCountMismatch message — the "" placeholder is never read
    # on that path. On the matching-count path every index is a real
    # position.
    return [
        DiePick.from_xy(
            sub_positions[i] if i < len(sub_positions) else "",
            wafer_ring,
            sel["x"],
            sel["y"],
            bin=str(sel.get("bin", "1")),
        )
        for i, sel in enumerate(selections)
    ]


@app.post("/api/generate")
def api_generate():
    data = request.get_json(force=True)
    try:
        template_positions = data.get("template_positions")
        if template_positions:
            # 複製既有.strate為範本 path: reuse the source file's own
            # position order verbatim instead of re-deriving it from
            # convention/machine_type — see blank_from_positions()'s
            # docstring for why that's the safer choice here.
            blank = blank_from_positions(
                assy_lot=data["assy_lot"],
                mapping_lot=data["mapping_lot"],
                eqpid=data["eqpid"],
                oper=data["oper"],
                substrate_id=data["substrate_id"],
                substrate_row=int(data["substrate_row"]),
                substrate_column=int(data["substrate_column"]),
                substrate_block=int(data["substrate_block"]),
                notch=data.get("notch", ""),
                ref=data.get("ref", ""),
                positions=template_positions,
            )
        else:
            blank = _blank_from_header(data)
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400

    wafer_ring = data.get("wafer_ring", "")
    # Positions the operator explicitly marked "不上片" (no die at this
    # substrate site) are dropped from the fillable list up front — this
    # reuses assign_dies()/assign_two_layers()'s existing "unfilled
    # positions are simply absent from DIE_INFO" behavior (see
    # bingomap/CLAUDE.md), it just shrinks what counts as "unfilled" and
    # the required pick count to match.
    skip_positions = set(data.get("skip_positions") or [])
    sub_positions = [d.sub_pos for d in blank.die_info if d.sub_pos not in skip_positions]
    target_qty = len(sub_positions)

    try:
        start_time = datetime.fromisoformat(data.get("start_time"))
    except (TypeError, ValueError):
        return jsonify({"error": "start_time must be an ISO 8601 timestamp"}), 400
    interval_seconds = int(data.get("interval_seconds", 2))

    try:
        layers_data = data.get("layers")
        if layers_data:
            # N-layer (一次上N顆) path — layers_data[i] is layer i+1's
            # selections; the LAST layer is the current/topmost one and
            # goes into DIE_INFO, everything else into DIE_INFO_OTHER_LAYER
            # (see assign_layers() docstring / bingomap/CLAUDE.md).
            layer_picks = [
                _build_picks(layer_selections, sub_positions, wafer_ring)
                for layer_selections in layers_data
            ]
            filled = assign_layers(
                blank,
                layer_picks,
                start_time=start_time,
                interval_seconds=interval_seconds,
                expected_qty=target_qty,
            )
        else:
            picks = _build_picks(data.get("selections", []), sub_positions, wafer_ring)
            filled = assign_dies(
                blank,
                picks,
                start_time=start_time,
                interval_seconds=interval_seconds,
                expected_qty=target_qty,
            )
    except DieCountMismatch as exc:
        return jsonify({"error": str(exc)}), 409
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    timestamp = start_time.strftime("%Y%m%d%H%M%S")
    filename = filled.filename(timestamp)
    content = filled.to_text()
    return Response(
        content,
        mimetype="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/mispick")
def mispick_page():
    return render_template("mispick.html", active_page="mispick")


_MISPICK_CSV_HEADER = [
    "source_file", "substrate_id", "layer", "action_no", "decision",
    "output_block", "output_coord", "tx", "ty", "fx", "fy",
    "nominal_map_x", "nominal_map_y", "nominal_bin",
    "actual_map_x", "actual_map_y", "actual_bin", "wafer_ring",
]


def _csv_text(rows: list[list]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    for row in rows:
        writer.writerow(row)
    return "﻿" + buf.getvalue()  # BOM so Excel opens it as UTF-8, not Big5


@app.post("/api/mispick/analyze")
def api_mispick_analyze():
    """誤吸偏移／BIN點除: given a known machine pick offset, the original
    wafer bin map (FRM), and one or more already-produced STRATE files,
    figure out which placed dies actually landed on a bad wafer BIN and
    need to be point-removed. See bingomap/mispick_analysis.py —
    machine_type="DB" (the default, this project's real machine type,
    confirmed against a real DB case 2026/08/17) or "ESEC" (ported from a
    reference tool, NOTCH=270 only, not this project's own machine)."""
    data = request.get_json(force=True)

    wafer_ring = (data.get("wafer_ring") or "").strip()
    if not wafer_ring:
        return jsonify({"error": "請輸入要比對的完整Wafer ID"}), 400

    machine_type = data.get("machine_type", "DB")

    try:
        offset = make_offset(data.get("offset_axis", "X"), int(data.get("offset_value", 0)))
    except (TypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400

    good_bins = parse_bin_set(data.get("good_bins", ""), default="1")
    ng_bins = parse_bin_set(data.get("ng_bins", ""), default="7,9")
    review_bins = parse_bin_set(data.get("review_bins", ""), default="2")

    frm_info = data.get("frm") or {}
    lot_no = (frm_info.get("lot_no") or "").strip()
    barcode_id = (frm_info.get("barcode_id") or "").strip()
    frm_root = (frm_info.get("frm_path") or DEFAULT_FRM_PATH).strip()
    if not lot_no or not barcode_id:
        return jsonify({"error": "請輸入原始wafer MAP的FRM Lot No跟Barcode ID"}), 400
    try:
        path = frm_file_path(frm_root, lot_no, barcode_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    fs_path = path.replace("\\", os.sep)
    if not os.path.exists(fs_path):
        return jsonify({"error": f"找不到原始wafer MAP檔案：{path}"}), 404
    try:
        with open(fs_path, "rb") as f:
            frm = parse_frm(f.read())
    except FrmFormatError as exc:
        return jsonify({"error": f"原始wafer MAP格式解析失敗：{exc}"}), 422
    wafer_map = frm_to_wafer_bin_map(frm)

    strate_files = data.get("strate_files") or []
    if not strate_files:
        return jsonify({"error": "請至少上傳一份STRATE檔案"}), 400

    substrates_out = []
    csv_rows = [_MISPICK_CSV_HEADER]
    for item in strate_files:
        name = item.get("name", "")
        text = item.get("text", "")
        try:
            substrate = StrateFile.parse(text)
        except StrateFormatError as exc:
            substrates_out.append({"name": name, "substrate_id": None, "error": f"STRATE格式解析失敗：{exc}"})
            continue
        try:
            result = analyze_substrate(
                substrate,
                wafer_map,
                wafer_ring=wafer_ring,
                offset=offset,
                good_bins=good_bins,
                ng_bins=ng_bins,
                review_bins=review_bins,
                machine_type=machine_type,
            )
        except (UnsupportedNotchError, InvalidGeometryError, UnknownMachineTypeError) as exc:
            substrates_out.append({"name": name, "substrate_id": substrate.substrate_id, "error": str(exc)})
            continue

        summary = {"force_delete": 0, "review": 0, "anomaly": 0, "ok": 0, "other": 0}
        action_rows_out = []
        grid_cells_out = []
        for row in result.rows:
            if row.decision == DECISION_FORCE_DELETE:
                summary["force_delete"] += 1
            elif row.decision == DECISION_REVIEW:
                summary["review"] += 1
            elif row.decision == DECISION_ANOMALY:
                summary["anomaly"] += 1
            elif row.decision == DECISION_OK:
                summary["ok"] += 1
            else:
                summary["other"] += 1

            nominal_x, nominal_y = row.nominal_map_xy or ("", "")
            actual_x, actual_y = row.actual_map_xy or ("", "")
            csv_rows.append(
                [
                    name, substrate.substrate_id, row.layer, row.action_no or "", row.decision,
                    row.output_block or "", output_coord(row), row.tx, row.ty, row.fx, row.fy,
                    nominal_x, nominal_y, row.nominal_bin or "",
                    actual_x, actual_y, row.actual_bin or "", row.source_die.wafer_ring,
                ]
            )
            if row.action_no is not None:
                action_rows_out.append(
                    {
                        "action_no": row.action_no,
                        "decision": row.decision,
                        "layer": row.layer,
                        "output_block": row.output_block,
                        "output_coord": output_coord(row),
                        "tx": row.tx,
                        "ty": row.ty,
                        "actual_bin": row.actual_bin,
                    }
                )
            # Every placed die (not just the ones needing an action) — for
            # drawing the substrate's own BINGO MAP grid with force-delete
            # positions outlined in red, per the user's explicit ask
            # 2026/08/18 ("要反紅框讓我知道"). tx/ty are the substrate's
            # own DIE_INFO sub_pos, not the machine-type output-mirrored
            # position — fine since this project's real machine type (DB)
            # is an identity mapping between the two (see output_position()
            # in mispick_analysis.py); only ESEC substrates would visually
            # differ from the existing "output_coord" table column, and
            # ESEC already carries its own prominent warning on this page.
            if row.tx is not None and row.ty is not None:
                grid_cells_out.append(
                    {"tx": row.tx, "ty": row.ty, "decision": row.decision, "layer": row.layer}
                )
        action_rows_out.sort(key=lambda r: r["action_no"])

        substrates_out.append(
            {
                "name": name,
                "substrate_id": substrate.substrate_id,
                "error": None,
                "summary": summary,
                "excluded_count": len(result.excluded),
                "action_rows": action_rows_out,
                "substrate_column": substrate.substrate_column,
                "substrate_row": substrate.substrate_row,
                "grid_cells": grid_cells_out,
            }
        )

    return jsonify(
        {
            "wafer": {
                "columns": frm.col,
                "rows": frm.row,
                "lot_no": frm.lot_no,
                "wafer_id": frm.wafer_id,
                "cells": [{"x": x, "y": y, "bin": str(bin_kind)} for (x, y), bin_kind in frm.die_map.items()],
            },
            "substrates": substrates_out,
            "csv": _csv_text(csv_rows),
        }
    )


@app.get("/crack")
def crack_page():
    return render_template("crack.html", active_page="crack")


def _crack_doc_payload(doc_index, name, substrate, candidates):
    cells = [
        {
            "key": c.key,
            "tx": c.tx,
            "ty": c.ty,
            "output_x": c.output_xy[0],
            "output_y": c.output_xy[1],
            "output_block": c.output_block,
            "output_coord": crack_output_coord(c),
            "wafer_id": c.wafer_id,
            "fx": c.fx,
            "fy": c.fy,
        }
        for c in candidates
        if c.doc_index == doc_index
    ]
    return {
        "doc_index": doc_index,
        "name": name,
        "substrate_id": substrate.substrate_id,
        "row": substrate.substrate_row,
        "column": substrate.substrate_column,
        "block": substrate.substrate_block,
        "cells": cells,
    }


@app.post("/api/crack/analyze")
def api_crack_analyze():
    """Crack位置回推: pool DIE_INFO rows sharing a wafer ID across every
    uploaded STRATE, let the operator mark crack positions on each
    substrate's own grid, and reconstruct where those marks sit relative to
    each other on the source wafer (local scatter only — see
    bingomap/crack_recovery.py for why this is explicitly not a true
    absolute wafer position). Stateless like the rest of this app: every
    call re-parses the uploaded STRATE text and re-derives everything from
    `marked_keys`, so there's nothing to keep in sync server-side."""
    data = request.get_json(force=True)
    strate_files = data.get("strate_files") or []
    if not strate_files:
        return jsonify({"error": "請至少上傳一份STRATE檔案"}), 400

    machine_type = data.get("machine_type", "DB")

    docs = []
    for item in strate_files:
        name = item.get("name", "")
        text = item.get("text", "")
        try:
            substrate = StrateFile.parse(text)
        except StrateFormatError as exc:
            return jsonify({"error": f"{name} 格式解析失敗：{exc}"}), 422
        docs.append((name, substrate))

    try:
        session = build_session(docs, machine_type=machine_type)
    except (InvalidGeometryError, MissingNotchError, UnknownMachineTypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 422

    marked_keys = [k for k in (data.get("marked_keys") or []) if k in session.by_key]
    wafer_ids = session.wafer_ids()

    focus_wafer_id = data.get("focus_wafer_id") or ""
    if focus_wafer_id not in wafer_ids:
        marked = session.marked(marked_keys)
        focus_wafer_id = marked[-1].wafer_id if marked else session.candidates[0].wafer_id

    docs_out = [_crack_doc_payload(i, name, substrate, session.candidates) for i, (name, substrate) in enumerate(docs)]

    rng, notch, points = wafer_scatter(session, focus_wafer_id, marked_keys, machine_type=machine_type)
    scatter = {
        "range": {"min_x": rng.min_x, "max_x": rng.max_x, "min_y": rng.min_y, "max_y": rng.max_y},
        "notch": notch,
        "points": [
            {"key": p.key, "x": p.x, "y": p.y, "is_crack": p.is_crack, "crack_no": p.crack_no} for p in points
        ],
    }

    crack_table = [
        {
            "crack_no": i,
            "key": c.key,
            "substrate_id": c.substrate_id,
            "source": c.doc_name,
            "output_block": c.output_block,
            "output_coord": crack_output_coord(c),
            "tx": c.tx,
            "ty": c.ty,
            "wafer_id": c.wafer_id,
            "fx": c.fx,
            "fy": c.fy,
            "notch": c.notch,
        }
        for i, c in enumerate(session.marked(marked_keys), start=1)
    ]

    return jsonify(
        {
            "docs": docs_out,
            "wafer_ids": wafer_ids,
            "focus_wafer_id": focus_wafer_id,
            "scatter": scatter,
            "crack_table": crack_table,
            "csv": _csv_text(crack_csv_rows(session, marked_keys, machine_type=machine_type)),
        }
    )


@app.get("/strate-xml")
def strate_xml_page():
    return render_template("strate_xml.html", active_page="strate_xml")


def _decode_uploaded_log(data: dict) -> tuple[str | None, tuple[dict, int] | None]:
    """Shared by both /api/strate_xml endpoints: base64-decode the
    uploaded log bytes and run them through decode_secs_log()'s UTF-16LE
    detection (see bingomap/secs_log.py — the real log this was built
    from has no BOM). Returns (text, None) on success, or
    (None, (error_response_dict, status_code)) on failure."""
    log_b64 = data.get("log_base64", "")
    if not log_b64:
        return None, ({"error": "請提供SECS log檔案"}, 400)
    try:
        raw = base64.b64decode(log_b64)
    except (ValueError, binascii.Error) as exc:
        return None, ({"error": f"檔案內容解碼失敗：{exc}"}, 400)
    return decode_secs_log(raw), None


def _substrate_die_positions(sf: StrateFile) -> list[dict]:
    """Every wafer coordinate this substrate consumed, across DIE_INFO and
    DIE_INFO_OTHER_LAYER — used to draw this substrate's footprint on its
    source wafer's bin map (see the wafer-map overlay in strate_xml.js).
    `d.wafer_xy` is already col:row here — a StrateMap's DIE_INFO wafer_xy
    is natively in that format already, no transform needed (2026/08/26
    correction, see bingomap/secs_log.py's module docstring) — x/y below
    line up directly with `_wafer_map_summary()`'s cells, both col:row."""
    positions = []
    for d in sf.die_info + sf.other_layer_die_info:
        x_str, _, y_str = d.wafer_xy.partition(":")
        try:
            positions.append({"x": int(x_str), "y": int(y_str)})
        except ValueError:
            continue
    return positions


def _substrate_summary(sf: StrateFile, index: int) -> dict:
    last_ts = max((d.timestamp for d in sf.die_info), default="")
    wafer_ring = sf.die_info[0].wafer_ring if sf.die_info else ""
    base_name = sf.substrate_id or f"substrate_{index + 1}"
    return {
        "index": index,
        "substrate_id": sf.substrate_id,
        "wafer_ring": wafer_ring,
        "eqpid": sf.eqpid,
        "num_dies": len(sf.die_info),
        "num_other_layer_dies": len(sf.other_layer_die_info),
        "total_bond_die_qty": sf.total_bond_die_qty,
        "good_die": sf.good_die,
        "last_timestamp": last_ts,
        "filename": f"{base_name}_{last_ts}.strate" if last_ts else f"{base_name}.strate",
        "text": sf.to_text(),
        "die_positions": _substrate_die_positions(sf),
    }


def _wafer_map_summary(wm, index: int) -> dict:
    cells = wm.wafer_map.cells
    paste_text = "\n".join(f"{x},{y},{bin_}" for (x, y), bin_ in sorted(cells.items()))
    return {
        "index": index,
        "frame_id": wm.frame_id,
        "wafer_id": wm.wafer_id,
        "columns": wm.wafer_map.columns,
        "rows": wm.wafer_map.rows,
        "num_cells": len(cells),
        "cells": [{"x": x, "y": y, "bin": bin_} for (x, y), bin_ in cells.items()],
        "paste_text": paste_text,
    }


@app.post("/api/strate_xml/extract")
def api_strate_xml_extract():
    """④ STRATE補檔 XML合併: pull already-complete substrate data
    (StrateMap events) and wafer bin maps (WaferStart events' BinList)
    straight out of a machine's SECS/AFC transaction log — see
    bingomap/secs_log.py's module docstring for how each was verified
    against a real log. No coordinate re-picking needed for the substrate
    side; ASSY_LOT/MAPPING_LOT/OPER aren't in this log at all, so they
    come back blank for the operator to fill in by hand."""
    data = request.get_json(force=True)
    text, err = _decode_uploaded_log(data)
    if err:
        return jsonify(err[0]), err[1]

    strate_files = extract_strate_files(text)
    wafer_maps = extract_wafer_maps(text)

    return jsonify(
        {
            "substrates": [_substrate_summary(sf, i) for i, sf in enumerate(strate_files)],
            "wafer_maps": [_wafer_map_summary(wm, i) for i, wm in enumerate(wafer_maps)],
        }
    )


@app.post("/api/strate_xml/download_zip")
def api_strate_xml_download_zip():
    """Batch download: every StrateMap event in the uploaded log, each as
    its own .strate file, bundled into one zip — re-parses the same
    uploaded log rather than trusting client-supplied file contents."""
    data = request.get_json(force=True)
    text, err = _decode_uploaded_log(data)
    if err:
        return jsonify(err[0]), err[1]

    strate_files = extract_strate_files(text)
    if not strate_files:
        return jsonify({"error": "這份log裡沒有找到StrateMap(基板)資料"}), 400

    buf = io.BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, sf in enumerate(strate_files):
            summary = _substrate_summary(sf, i)
            name = summary["filename"]
            n = 2
            while name in used_names:
                name = f"{summary['filename'][:-7]}_{n}.strate"  # strip ".strate", renumber
                n += 1
            used_names.add(name)
            zf.writestr(name, summary["text"])
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": 'attachment; filename="strate_xml_extract.zip"'},
    )


@app.get("/secs-params")
def secs_params_page():
    return render_template("secs_params.html", active_page="secs_params")


@app.get("/api/secs_params/baseline")
def api_secs_params_baseline():
    """The fixed 199-parameter baseline (see SECS_PARAMS_BASELINE_PATH's
    comment) — the page loads this immediately on open, no log upload
    needed. "上傳SECS Log" below stays available separately, for checking
    a *different* log's own snapshot against this baseline."""
    baseline = _load_secs_params_baseline()
    return jsonify(
        {
            "pp_id": baseline["pp_id"],
            "mdln": baseline["mdln"],
            "softrev": baseline["softrev"],
            "source_tid": baseline["source_tid"],
            "params": baseline["params"],
        }
    )


@app.post("/api/secs_params/extract")
def api_secs_params_extract():
    """List every SECS "formatted process program" parameter snapshot
    (S7F25FormattedPPRequest Reply) found in an uploaded SECS/AFC log —
    see bingomap/secs_params.py. Excel-comparison is not implemented yet;
    this is the "list what's actually in the log" half only."""
    data = request.get_json(force=True)
    text, err = _decode_uploaded_log(data)
    if err:
        return jsonify(err[0]), err[1]

    try:
        snapshots = extract_pp_param_snapshots(text)
    except SecsParamsFormatError as exc:
        return jsonify({"error": f"參數格式解析失敗：{exc}"}), 422

    return jsonify(
        {
            "snapshots": [
                {
                    "index": i,
                    "pp_id": s.pp_id,
                    "mdln": s.mdln,
                    "softrev": s.softrev,
                    "tid": s.tid,
                    "params": [
                        {
                            "ccode": p.ccode,
                            "name": p.name,
                            "unit": p.unit,
                            "format": p.format,
                            "value": p.value,
                            "min": p.min,
                            "max": p.max,
                        }
                        for p in s.params
                    ],
                }
                for i, s in enumerate(snapshots)
            ]
        }
    )


class ExcelChecklistFormatError(ValueError):
    """Raised when the uploaded checklist .xlsx doesn't have the expected
    4-column header — see _parse_excel_checklist()."""


_CHECKLIST_EXPECTED_HEADER = ["Item", "Name", "ID number", "ID name"]


def _checklist_cell_str(value) -> str:
    return "" if value is None else str(value).strip()


def _checklist_ccode_str(value) -> str:
    # openpyxl reads a numeric CCODE cell back as int/float, not str — the
    # baseline's own ccode values are plain digit strings (see
    # bingomap/data/secs_params_baseline.json), so normalize here rather
    # than at every comparison call site.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_excel_checklist(file_bytes: bytes) -> list[ChecklistRow]:
    """Parse a user-uploaded target-parameter checklist .xlsx into
    ChecklistRow objects — see bingomap/secs_params.py's module comment on
    why CCODE ("ID number") is the match key, confirmed 2026/08/20 against
    a real checklist (185/199 baseline CCODEs found by exact match).
    Strict on the header (first sheet, first row must be exactly
    Item/Name/ID number/ID name) — this project's convention is to error
    out on an unexpected format rather than guess a column mapping."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises several exception types for a bad file
        raise ExcelChecklistFormatError(f"Excel檔案無法讀取：{exc}") from exc

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ExcelChecklistFormatError("Excel檔案是空的")

    header = [_checklist_cell_str(c) for c in rows[0][:4]]
    if header != _CHECKLIST_EXPECTED_HEADER:
        raise ExcelChecklistFormatError(
            f"第一列欄位標題預期是 {_CHECKLIST_EXPECTED_HEADER}，實際讀到 {header}"
            "——請確認是不是同一份格式的Excel，欄位對應不會用猜的"
        )

    checklist_rows: list[ChecklistRow] = []
    current_category = ""
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        item, name, id_number, id_name = (list(r) + [None, None, None, None])[:4]
        category = _checklist_cell_str(item)
        if category:
            current_category = category
        if id_number is None or name is None:
            continue  # no CCODE or no name on this row — nothing to match, skip
        checklist_rows.append(
            ChecklistRow(
                category=current_category,
                name=_checklist_cell_str(name),
                ccode=_checklist_ccode_str(id_number),
                id_name=_checklist_cell_str(id_name),
            )
        )
    return checklist_rows


def _checklist_row_dict(row: ChecklistRow) -> dict:
    return {"category": row.category, "name": row.name, "ccode": row.ccode, "id_name": row.id_name}


@app.post("/api/secs_params/compare_excel")
def api_secs_params_compare_excel():
    """Check-list式比對：把上傳的目標參數Excel(CCODE為鍵)跟固定的199組
    基準參數清單比對，分成「兩邊都有」「機台有Excel沒有」「Excel有機台
    沒有(要新增的)」三類——見bingomap/secs_params.py的compare_checklist()。"""
    data = request.get_json(force=True)
    file_b64 = data.get("excel_base64")
    if not file_b64:
        return jsonify({"error": "缺少Excel檔案內容"}), 400
    try:
        file_bytes = base64.b64decode(file_b64)
    except (ValueError, binascii.Error) as exc:
        return jsonify({"error": f"檔案內容解碼失敗：{exc}"}), 400

    try:
        checklist_rows = _parse_excel_checklist(file_bytes)
    except ExcelChecklistFormatError as exc:
        return jsonify({"error": str(exc)}), 422

    if not checklist_rows:
        return jsonify({"error": "Excel檔案裡沒有解析到任何一列有效的參數(ID number/Name都要有值)"}), 422

    baseline = _load_secs_params_baseline()
    baseline_params = [
        SecsParam(
            ccode=p["ccode"], name=p["name"], unit=p["unit"], format=p["format"],
            value=p["value"], min=p["min"], max=p["max"],
        )
        for p in baseline["params"]
    ]

    comparison = compare_checklist(baseline_params, checklist_rows)

    return jsonify(
        {
            "counts": {
                "checklist_total_rows": len(checklist_rows),
                "checklist_unique_ccodes": len(comparison.matched) + len(comparison.checklist_only),
                "baseline_total": len(baseline_params),
                "matched": len(comparison.matched),
                "machine_only": len(comparison.machine_only),
                "checklist_only": len(comparison.checklist_only),
                "duplicate_ccode_groups": len(comparison.duplicate_ccodes),
            },
            "matched": sorted(comparison.matched, key=lambda m: m["ccode"]),
            "machine_only": sorted(comparison.machine_only, key=lambda m: m["ccode"]),
            "checklist_only": [
                _checklist_row_dict(r) for r in sorted(comparison.checklist_only, key=lambda r: r.ccode)
            ],
            "duplicate_ccodes": [
                {"ccode": d["ccode"], "rows": [_checklist_row_dict(r) for r in d["rows"]]}
                for d in sorted(comparison.duplicate_ccodes, key=lambda d: d["ccode"])
            ],
        }
    )


_SECS_PARAMS_HEADER = ["CCODE", "名稱", "單位", "格式", "數值", "下限", "上限"]


def _snapshot_to_dict(snap) -> dict:
    """PPParamSnapshot (from a freshly-parsed log) -> the same plain-dict
    shape the baseline JSON file already uses, so the TXT/Excel builders
    below only need to handle one shape."""
    return {
        "pp_id": snap.pp_id,
        "mdln": snap.mdln,
        "softrev": snap.softrev,
        "tid": snap.tid,
        "params": [
            {"ccode": p.ccode, "name": p.name, "unit": p.unit, "format": p.format, "value": p.value, "min": p.min, "max": p.max}
            for p in snap.params
        ],
    }


def _param_row(p: dict) -> list[str]:
    return [p["ccode"], p["name"], p["unit"], p["format"], p["value"], p["min"], p["max"]]


def _extract_snapshots_or_error(data: dict):
    """Shared by the log-based export endpoints below: re-decode+re-parse
    the uploaded log server-side (same principle as
    /api/strate_xml/download_zip — never trust client-supplied result data
    for a download, re-derive it from the source). Returns
    (snapshot_dicts, None) or (None, (json, status))."""
    text, err = _decode_uploaded_log(data)
    if err:
        return None, err
    try:
        snapshots = extract_pp_param_snapshots(text)
    except SecsParamsFormatError as exc:
        return None, ({"error": f"參數格式解析失敗：{exc}"}, 422)
    if not snapshots:
        return None, ({"error": "這份log裡沒有找到S7F25FormattedPPRequest參數快照"}, 400)
    return [_snapshot_to_dict(s) for s in snapshots], None


def _txt_for_snapshots(snapshot_dicts: list[dict]) -> bytes:
    lines = []
    for i, snap in enumerate(snapshot_dicts):
        tid_note = f"\tTID={snap['tid']}" if snap.get("tid") else ""
        lines.append(
            f"# 快照{i + 1}\tPP_ID={snap['pp_id']}\tMDLN={snap['mdln']}\tSOFTREV={snap['softrev']}"
            f"{tid_note}\t參數數量={len(snap['params'])}"
        )
        lines.append("\t".join(_SECS_PARAMS_HEADER))
        for p in snap["params"]:
            lines.append("\t".join(str(v) for v in _param_row(p)))
        lines.append("")
    content = "\r\n".join(lines) + "\r\n"
    return content.encode("utf-8-sig")  # BOM so Excel/記事本 opens it as UTF-8, not Big5


@app.post("/api/secs_params/download_txt")
def api_secs_params_download_txt():
    data = request.get_json(force=True)
    snapshots, err = _extract_snapshots_or_error(data)
    if err:
        return jsonify(err[0]), err[1]

    return Response(
        _txt_for_snapshots(snapshots),
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="secs_params.txt"'},
    )


@app.get("/api/secs_params/baseline/download_txt")
def api_secs_params_baseline_download_txt():
    baseline = _load_secs_params_baseline()
    snap = {
        "pp_id": baseline["pp_id"], "mdln": baseline["mdln"], "softrev": baseline["softrev"],
        "tid": baseline["source_tid"], "params": baseline["params"],
    }
    return Response(
        _txt_for_snapshots([snap]),
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="secs_params_baseline.txt"'},
    )


def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel sheet names: max 31 chars, can't contain \ / * ? : [ ]
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "snapshot"
    cleaned = cleaned[:31]
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _excel_for_snapshots(snapshot_dicts: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    for i, snap in enumerate(snapshot_dicts):
        # TID first: Excel's 31-char sheet-name limit would otherwise often
        # truncate off the TID entirely when multiple snapshots share the
        # same (long) PP_ID, leaving indistinguishable names that only
        # differ by the dedup suffix _safe_sheet_name() adds — found
        # 2026/08/19 with the real log's two RECIPE@AEU132X2C001A-2070
        # captures, which both truncated to the same 31 chars.
        tid = snap.get("tid") or ""
        tid_label = f"TID{tid}" if tid else f"snapshot{i + 1}"
        sheet_name = _safe_sheet_name(f"{tid_label}_{snap['pp_id']}", used_names)
        ws = wb.create_sheet(sheet_name)
        ws.append([f"PP_ID={snap['pp_id']}", f"MDLN={snap['mdln']}", f"SOFTREV={snap['softrev']}", f"TID={tid}"])
        ws.append(_SECS_PARAMS_HEADER)
        rows = [_param_row(p) for p in snap["params"]]
        for row in rows:
            ws.append(row)
        for col_idx, header in enumerate(_SECS_PARAMS_HEADER, start=1):
            width = len(header)
            for row in rows:
                width = max(width, len(str(row[col_idx - 1])))
            ws.column_dimensions[chr(64 + col_idx)].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.post("/api/secs_params/download_excel")
def api_secs_params_download_excel():
    data = request.get_json(force=True)
    snapshots, err = _extract_snapshots_or_error(data)
    if err:
        return jsonify(err[0]), err[1]

    return Response(
        _excel_for_snapshots(snapshots),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="secs_params.xlsx"'},
    )


@app.get("/api/secs_params/baseline/download_excel")
def api_secs_params_baseline_download_excel():
    baseline = _load_secs_params_baseline()
    snap = {
        "pp_id": baseline["pp_id"], "mdln": baseline["mdln"], "softrev": baseline["softrev"],
        "tid": baseline["source_tid"], "params": baseline["params"],
    }
    return Response(
        _excel_for_snapshots([snap]),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="secs_params_baseline.xlsx"'},
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
