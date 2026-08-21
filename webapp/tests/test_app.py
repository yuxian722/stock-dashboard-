from pathlib import Path

import pytest

from webapp.app import app

REAL_FRM_FIXTURE = (
    Path(__file__).parent.parent.parent / "bingomap" / "tests" / "fixtures" / "8P065800A1_T3_DA62.frm"
)
REAL_FRM_WITH_REF_POINT_FIXTURE = (
    Path(__file__).parent.parent.parent / "bingomap" / "tests" / "fixtures" / "8P964000_K8_4E13.frm"
)
REAL_STRATE_FIXTURE = (
    Path(__file__).parent.parent.parent
    / "bingomap"
    / "tests"
    / "fixtures"
    / "2070_V27NVJH_Z281226C_20260812221959.strate"
)
REAL_EIGHT_LAYER_STRATE_FIXTURE = (
    Path(__file__).parent.parent.parent
    / "bingomap"
    / "tests"
    / "fixtures"
    / "2070_V25NVDY_F2006908_20260702203138.strate"
)
REAL_SECS_LOG_FIXTURE = (
    Path(__file__).parent.parent.parent / "bingomap" / "tests" / "fixtures" / "secs_log_sample.log"
)
REAL_SECS_PARAMS_FIXTURE = (
    Path(__file__).parent.parent.parent / "bingomap" / "tests" / "fixtures" / "secs_params_sample.log"
)

BASE_HEADER = dict(
    assy_lot="V27NVJH",
    mapping_lot="S7MJS",
    eqpid="BAB12",
    oper="2070",
    substrate_id="Z281226C",
    substrate_row=4,
    substrate_column=20,
    substrate_block=2,
    notch="180",
    ref="-72,340",
    convention="EPOXY",
)


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


def test_index_page_loads(client):
    res = client.get("/")
    assert res.status_code == 200
    assert b"BINGO MAP" in res.data


def test_api_blank_esec_machine_type_starts_at_last_position(client):
    res = client.post("/api/blank", json={**BASE_HEADER, "machine_type": "ESEC"})
    assert res.status_code == 200
    positions = res.get_json()["positions"]
    assert positions[0] == "19:3"  # COLUMN-1:ROW-1 for ROW=4, COLUMN=20


def test_api_blank_defaults_to_db_machine_type(client):
    res = client.post("/api/blank", json=BASE_HEADER)  # no machine_type key at all
    positions = res.get_json()["positions"]
    assert positions[0] == "0:0"


def test_api_blank_returns_positions_and_qty(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    assert res.status_code == 200
    data = res.get_json()
    assert data["total_qty"] == 80
    assert len(data["positions"]) == 80
    assert data["positions"][0] == "0:0"


def test_api_blank_missing_field_returns_400(client):
    bad = {k: v for k, v in BASE_HEADER.items() if k != "substrate_row"}
    res = client.post("/api/blank", json=bad)
    assert res.status_code == 400


def test_api_generate_success_downloads_strate_file(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    positions = res.get_json()["positions"]

    selections = [{"x": 23, "y": 195 + i, "bin": "1"} for i in range(80)]
    payload = {
        **BASE_HEADER,
        "wafer_ring": "A27572",
        "start_time": "2026-08-12T22:16:33",
        "interval_seconds": 2,
        "selections": selections,
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    assert "attachment" in res.headers["Content-Disposition"]
    assert "2070_V27NVJH_Z281226C_20260812221633.strate" in res.headers["Content-Disposition"]

    text = res.get_data(as_text=True)
    assert "ASSY_LOT=V27NVJH" in text
    assert "TOTAL_BOND_DIE_QTY=80" in text
    assert text.count("\r\n") > 80
    assert positions[0] in text


def test_api_generate_skip_positions_reduces_required_qty(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    positions = res.get_json()["positions"]
    skip_positions = positions[:2]  # mark 2 substrate sites "不上片"

    selections = [{"x": 23, "y": 195 + i, "bin": "1"} for i in range(78)]  # 80 - 2 skipped
    payload = {
        **BASE_HEADER,
        "wafer_ring": "A27572",
        "start_time": "2026-08-12T22:16:33",
        "interval_seconds": 2,
        "selections": selections,
        "skip_positions": skip_positions,
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    text = res.get_data(as_text=True)
    assert "TOTAL_BOND_DIE_QTY=78" in text
    for pos in skip_positions:
        assert f",{pos},1,0,0," not in text  # never written as a DIE_INFO row


def test_api_generate_skip_positions_mismatch_uses_adjusted_target(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    positions = res.get_json()["positions"]
    skip_positions = positions[:2]

    payload = {
        **BASE_HEADER,
        "wafer_ring": "A27572",
        "start_time": "2026-08-12T22:16:33",
        "selections": [{"x": 1, "y": 1, "bin": "1"}],  # only 1, need 78 (80-2 skipped)
        "skip_positions": skip_positions,
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 409
    assert "需要 Die 數量78" in res.get_json()["error"]


def test_api_generate_quantity_mismatch_returns_dialog_wording(client):
    payload = {
        **BASE_HEADER,
        "wafer_ring": "A27572",
        "start_time": "2026-08-12T22:16:33",
        "interval_seconds": 2,
        "selections": [{"x": 1, "y": 1, "bin": "1"}],  # only 1, need 80
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 409
    data = res.get_json()
    assert "需要 Die 數量80" in data["error"]
    assert "已選擇數量1" in data["error"]
    assert "還需選擇79顆" in data["error"]


def test_api_generate_bad_start_time_returns_400(client):
    payload = {
        **BASE_HEADER,
        "wafer_ring": "A27572",
        "start_time": "not-a-date",
        "selections": [],
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 400


def _fake_frm_root(tmp_path, lot_no, barcode_id, fixture=REAL_FRM_FIXTURE):
    # Mirrors WaferCoordinate.exe's own layout: {root}\{LotNo}\{barcode[0:2]}\{barcode[2:6]}
    d = tmp_path / lot_no / barcode_id[0:2]
    d.mkdir(parents=True)
    (d / barcode_id[2:6]).write_bytes(fixture.read_bytes())
    return str(tmp_path)


def test_api_frm_loads_real_file_end_to_end(client, tmp_path):
    root = _fake_frm_root(tmp_path, "8P065800A1", "T3DA62")
    res = client.post(
        "/api/frm",
        json={"lot_no": "8P065800A1", "barcode_id": "T3DA62", "frm_path": root},
    )
    assert res.status_code == 200
    data = res.get_json()
    assert data["columns"] == 46
    assert data["rows"] == 56
    assert data["lot_no"] == "8P065800A1"
    assert data["wafer_id"] == "8P0658"
    assert data["wafer_type"] == "AW191"
    bins = [c["bin"] for c in data["cells"]]
    assert bins.count("1") == 1635
    assert bins.count("7") == 379


def test_api_frm_returns_reference_point_from_real_file(client, tmp_path):
    # Real file the user sent 2026/08/18 while tracking down what "T點" in
    # WaferCoordinate.exe's status bar actually is. This confirms the FRM
    # header's own reference_point_x/y field parses correctly (matches the
    # MAP INFORMATION AREA panel's bin1=1746/bin7=268) — NOT that this
    # field is T點 itself: a follow-up photo proved that guess wrong (the
    # field sits outside the wafer's actual die area, nowhere near the
    # real on-screen crosshair). See webapp/README.md and app.js's
    # refPointByPanel comment for how T點 is actually derived now.
    root = _fake_frm_root(tmp_path, "8P964000", "K84E13", fixture=REAL_FRM_WITH_REF_POINT_FIXTURE)
    res = client.post(
        "/api/frm",
        json={"lot_no": "8P964000", "barcode_id": "K84E13", "frm_path": root},
    )
    assert res.status_code == 200
    data = res.get_json()
    assert data["lot_no"] == "8P964000"
    assert data["wafer_id"] == "8P9640"
    assert data["reference_point_x"] == 5
    assert data["reference_point_y"] == 5
    bins = [c["bin"] for c in data["cells"]]
    assert bins.count("1") == 1746
    assert bins.count("7") == 268
    # The reference point itself isn't a real die.
    assert not any(c["x"] == 5 and c["y"] == 5 for c in data["cells"])
    # T點 itself (computed client-side in app.js, see refPointByPanel) is
    # (columns//2 - reference_point_x, 0) — confirmed against this same
    # file: the user pointed at the real T點 cell in WaferCoordinate.exe
    # and reported (18, 0). Pinning the inputs to that formula here so a
    # future change to reference_point_x/columns parsing can't silently
    # break it without a test noticing.
    t_point_x = data["columns"] // 2 - data["reference_point_x"]
    assert (t_point_x, 0) == (18, 0)


def test_api_frm_missing_file_returns_404_with_helpful_message(client, tmp_path):
    res = client.post(
        "/api/frm",
        json={"lot_no": "NOPE123", "barcode_id": "T3DA62", "frm_path": str(tmp_path)},
    )
    assert res.status_code == 404
    assert "找不到檔案" in res.get_json()["error"]


def test_api_frm_requires_lot_no_and_barcode(client):
    res = client.post("/api/frm", json={"lot_no": "", "barcode_id": ""})
    assert res.status_code == 400


def test_api_generate_two_layer_success(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    total_qty = res.get_json()["total_qty"]  # 80 for BASE_HEADER's 4x20

    # layers[-1] is the current/topmost layer -> f9=2 (DIE_INFO);
    # layers[0] -> f9=1 (DIE_INFO_OTHER_LAYER)
    other = [{"x": 21 - i, "y": 24, "bin": "1"} for i in range(total_qty)]
    primary = [{"x": 22 - i, "y": 24, "bin": "1"} for i in range(total_qty)]
    payload = {
        **BASE_HEADER,
        "wafer_ring": "I4F247",
        "start_time": "2026-08-12T16:15:21",
        "interval_seconds": 3,
        "layers": [other, primary],
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    text = res.get_data(as_text=True)
    assert "[DIE_INFO_OTHER_LAYER_BEG]" in text
    assert "[DIE_INFO_OTHER_LAYER_END]" in text
    assert text.count(",2\r\n") == total_qty
    assert text.count(",1\r\n") == total_qty


def test_api_generate_two_layer_mismatch_reports_which_side(client):
    res = client.post("/api/blank", json=BASE_HEADER)
    total_qty = res.get_json()["total_qty"]

    payload = {
        **BASE_HEADER,
        "wafer_ring": "I4F247",
        "start_time": "2026-08-12T16:15:21",
        "layers": [
            [{"x": 2, "y": 2, "bin": "1"}] * total_qty,
            [{"x": 1, "y": 1, "bin": "1"}],  # short (this is the last/current layer)
        ],
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 409
    data = res.get_json()
    assert f"需要 Die 數量{total_qty}" in data["error"]
    assert "已選擇數量1" in data["error"]


def test_api_generate_eight_layer_success(client):
    header = dict(BASE_HEADER, substrate_row=1, substrate_column=7, substrate_block=1)
    res = client.post("/api/blank", json=header)
    total_qty = res.get_json()["total_qty"]
    assert total_qty == 7

    layers = [[{"x": layer * 10 + i, "y": 1, "bin": "1"} for i in range(total_qty)] for layer in range(8)]
    payload = {
        **header,
        "wafer_ring": "B6844E",
        "start_time": "2026-07-02T20:26:40",
        "layers": layers,
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    text = res.get_data(as_text=True)
    assert "[DIE_INFO_OTHER_LAYER_BEG]" in text
    for f9 in range(1, 9):
        assert text.count(f",{f9}\r\n") == total_qty, f9


def _read_real_strate_text():
    with open(REAL_STRATE_FIXTURE, encoding="ascii", newline="") as f:
        return f.read()


def test_api_parse_strate_extracts_header_positions_and_picks(client):
    res = client.post("/api/parse_strate", json={"text": _read_real_strate_text()})
    assert res.status_code == 200
    data = res.get_json()
    assert data["assy_lot"] == "V27NVJH"
    assert data["substrate_id"] == "Z281226C"
    assert data["substrate_row"] == 4
    assert data["substrate_column"] == 20
    assert data["wafer_ring"] == "A27572"
    assert data["total_qty"] == 75
    assert len(data["positions"]) == 75
    assert data["positions"][0] == "0:0"
    assert data["positions"][-1] == "19:3"
    assert len(data["picks"]) == 75
    assert data["picks"][0] == {"x": 23, "y": 195, "bin": "1"}
    assert data["num_layers"] == 1
    assert data["layer_picks"] == [data["picks"]]


def test_api_parse_strate_splits_real_eight_layer_file_by_f9(client):
    with open(REAL_EIGHT_LAYER_STRATE_FIXTURE, encoding="ascii", newline="") as f:
        text = f.read()
    res = client.post("/api/parse_strate", json={"text": text})
    assert res.status_code == 200
    data = res.get_json()
    assert data["num_layers"] == 8
    assert len(data["layer_picks"]) == 8
    assert all(len(layer) == 7 for layer in data["layer_picks"])
    # layer_picks[-1] is the DIE_INFO (topmost/current) layer
    assert data["layer_picks"][-1] == data["picks"]
    # layer_picks[0] is f9=1's picks — first row's wafer_xy is "3:66"
    assert data["layer_picks"][0][0] == {"x": 3, "y": 66, "bin": "1"}


def test_api_parse_strate_rejects_malformed_text(client):
    res = client.post("/api/parse_strate", json={"text": "not a strate file"})
    assert res.status_code == 422


def test_api_parse_strate_rejects_empty_text(client):
    res = client.post("/api/parse_strate", json={"text": ""})
    assert res.status_code == 400


def test_api_generate_with_template_positions_bypasses_machine_type(client):
    parsed = client.post("/api/parse_strate", json={"text": _read_real_strate_text()}).get_json()
    payload = {
        "assy_lot": parsed["assy_lot"],
        "mapping_lot": parsed["mapping_lot"],
        "eqpid": parsed["eqpid"],
        "oper": parsed["oper"],
        "substrate_id": "Z999999Z",  # deliberately changed, as a real re-use would
        "substrate_row": parsed["substrate_row"],
        "substrate_column": parsed["substrate_column"],
        "substrate_block": parsed["substrate_block"],
        "notch": parsed["notch"],
        "ref": parsed["ref"],
        # No convention/machine_type at all — template_positions must win
        # regardless, proving this path never touches generate_blank().
        "wafer_ring": parsed["wafer_ring"],
        "start_time": "2026-08-14T09:00:00",
        "interval_seconds": 2,
        "template_positions": parsed["positions"],
        "selections": parsed["picks"],
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    text = res.get_data(as_text=True)
    assert "SUBSTRATE_ID=Z999999Z" in text
    assert "TOTAL_BOND_DIE_QTY=75" in text
    # exact same position order as the original real file, not a
    # DB/ESEC-regenerated one
    assert "0:0" in text
    assert "19:3" in text


def _secs_log_base64() -> str:
    import base64

    return base64.b64encode(REAL_SECS_LOG_FIXTURE.read_bytes()).decode("ascii")


def test_api_strate_xml_extract_real_log(client):
    res = client.post("/api/strate_xml/extract", json={"log_base64": _secs_log_base64()})
    assert res.status_code == 200
    data = res.get_json()

    assert len(data["substrates"]) == 2
    first = data["substrates"][0]
    assert first["substrate_id"] == "Z2570900444F"
    assert first["wafer_ring"] == "HD56BA"
    assert first["num_dies"] == 59
    assert first["total_bond_die_qty"] == 59
    assert first["good_die"] == 59
    # ASSY_LOT/OPER/MAPPING_LOT are blank — not present in this log at all.
    assert "ASSY_LOT=\r\n" in first["text"]
    assert "OPER=\r\n" in first["text"]
    assert "MAPPING_LOT=\r\n" in first["text"]
    assert "SUBSTRATE_ID=Z2570900444F" in first["text"]
    assert len(first["die_positions"]) == first["num_dies"] + first["num_other_layer_dies"]
    # 2026/08/21: wafer_xy is now normalized to col:row (x:y) — see
    # bingomap/secs_log.py's _swap_wafer_xy(); the log's raw DIE_INFO had
    # this as row:col ("10:42"), extraction now flips it to "42:10".
    assert first["die_positions"][0] == {"x": 42, "y": 10}

    assert len(data["wafer_maps"]) == 1
    wm = data["wafer_maps"][0]
    assert wm["frame_id"] == "HD66D5"
    assert wm["wafer_id"] == "P0264807-24"
    assert wm["columns"] == 46
    assert wm["rows"] == 24
    assert wm["num_cells"] > 0
    assert len(wm["cells"]) == wm["num_cells"]
    assert all({"x", "y", "bin"} <= c.keys() for c in wm["cells"])
    # paste_text must be directly usable in the main page's "x,y,bin" textarea.
    first_line = wm["paste_text"].splitlines()[0]
    assert len(first_line.split(",")) == 3


def test_api_strate_xml_extract_requires_log(client):
    res = client.post("/api/strate_xml/extract", json={})
    assert res.status_code == 400
    assert "log" in res.get_json()["error"] or "檔案" in res.get_json()["error"]


def test_api_strate_xml_extract_rejects_bad_base64(client):
    res = client.post("/api/strate_xml/extract", json={"log_base64": "not valid base64!!"})
    assert res.status_code == 400


def _secs_params_log_base64() -> str:
    import base64

    return base64.b64encode(REAL_SECS_PARAMS_FIXTURE.read_bytes()).decode("ascii")


def test_api_secs_params_extract_real_log(client):
    res = client.post("/api/secs_params/extract", json={"log_base64": _secs_params_log_base64()})
    assert res.status_code == 200
    data = res.get_json()

    assert len(data["snapshots"]) == 2
    snap = data["snapshots"][0]
    assert snap["pp_id"] == "RECIPE@AEU132X2C001A-2070"
    assert snap["mdln"] == "DB800"
    assert snap["softrev"] == "01.172/01"
    assert snap["tid"] == "58151"
    assert len(snap["params"]) == 10
    assert data["snapshots"][1]["tid"] == "58203"
    assert len(data["snapshots"][1]["params"]) == 3

    first = snap["params"][0]
    assert first["ccode"] == "285278212"
    assert first["name"] == "No. of blocks"
    assert first["format"] == "F8"
    assert first["value"] == "1"
    assert first["min"] == "0"
    assert first["max"] == "999"


def test_api_secs_params_extract_requires_log(client):
    res = client.post("/api/secs_params/extract", json={})
    assert res.status_code == 400


def test_api_secs_params_extract_rejects_bad_base64(client):
    res = client.post("/api/secs_params/extract", json={"log_base64": "not valid base64!!"})
    assert res.status_code == 400


def test_api_secs_params_download_txt_real_log(client):
    res = client.post("/api/secs_params/download_txt", json={"log_base64": _secs_params_log_base64()})
    assert res.status_code == 200
    assert "attachment" in res.headers["Content-Disposition"]
    text = res.get_data().decode("utf-8-sig")
    assert "PP_ID=RECIPE@AEU132X2C001A-2070" in text
    assert "TID=58151" in text
    assert "TID=58203" in text
    assert "285278212\tNo. of blocks" in text


def test_api_secs_params_download_txt_requires_log(client):
    res = client.post("/api/secs_params/download_txt", json={})
    assert res.status_code == 400


def test_api_secs_params_download_excel_real_log(client):
    import io

    import openpyxl

    res = client.post("/api/secs_params/download_excel", json={"log_base64": _secs_params_log_base64()})
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in res.headers["Content-Disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(res.get_data()))
    assert len(wb.sheetnames) == 2
    # Regression check for a real bug (2026/08/19): both snapshots share
    # the same long PP_ID, which used to make Excel's 31-char sheet-name
    # limit truncate away the only distinguishing part (TID) — see
    # app.py's _safe_sheet_name call site comment.
    assert "58151" in wb.sheetnames[0]
    assert "58203" in wb.sheetnames[1]
    ws = wb[wb.sheetnames[0]]
    assert "PP_ID=RECIPE@AEU132X2C001A-2070" in ws.cell(row=1, column=1).value
    assert ws.cell(row=2, column=1).value == "CCODE"
    assert ws.cell(row=3, column=1).value == "285278212"
    assert ws.cell(row=3, column=2).value == "No. of blocks"


def test_api_secs_params_download_excel_requires_log(client):
    res = client.post("/api/secs_params/download_excel", json={})
    assert res.status_code == 400


def test_api_secs_params_baseline(client):
    # 2026/08/19 ask: "把這頁改成這199格式化參數固定在裡面" — the page
    # loads this fixed baseline immediately, no log upload needed.
    res = client.get("/api/secs_params/baseline")
    assert res.status_code == 200
    data = res.get_json()
    assert data["pp_id"] == "RECIPE@AEU132X2C001A-2070"
    assert data["mdln"] == "DB800"
    assert data["softrev"] == "01.172/01"
    assert len(data["params"]) == 199
    first = data["params"][0]
    assert first["ccode"] == "285278212"
    assert first["name"] == "No. of blocks"
    ccodes = [p["ccode"] for p in data["params"]]
    assert len(ccodes) == len(set(ccodes))


def test_api_secs_params_baseline_download_txt(client):
    res = client.get("/api/secs_params/baseline/download_txt")
    assert res.status_code == 200
    assert "attachment" in res.headers["Content-Disposition"]
    text = res.get_data().decode("utf-8-sig")
    assert "PP_ID=RECIPE@AEU132X2C001A-2070" in text
    assert "285278212\tNo. of blocks" in text
    assert text.count("\n") > 199  # header + 199 param rows (+ blank line)


def test_api_secs_params_baseline_download_excel(client):
    import io

    import openpyxl

    res = client.get("/api/secs_params/baseline/download_excel")
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(res.get_data()))
    assert len(wb.sheetnames) == 1
    ws = wb[wb.sheetnames[0]]
    assert "PP_ID=RECIPE@AEU132X2C001A-2070" in ws.cell(row=1, column=1).value
    assert ws.cell(row=2, column=1).value == "CCODE"
    assert ws.max_row == 2 + 199


def _checklist_xlsx_bytes(rows) -> bytes:
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Item", "Name", "ID number", "ID name"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_api_secs_params_compare_excel_categorizes_matched_new_and_machine_only(client):
    import base64

    rows = [
        ("[Recipe]-[Group A]", "No. of blocks", 285278212, "DT_SOME_ID"),  # matches baseline exactly
        (None, "Brand new param", 999999999, "DT_NEW_ID"),  # not on the machine yet
    ]
    b64 = base64.b64encode(_checklist_xlsx_bytes(rows)).decode("ascii")
    res = client.post("/api/secs_params/compare_excel", json={"excel_base64": b64})
    assert res.status_code == 200
    data = res.get_json()

    assert data["counts"]["checklist_total_rows"] == 2
    assert data["counts"]["checklist_unique_ccodes"] == 2
    assert data["counts"]["matched"] == 1
    assert data["counts"]["checklist_only"] == 1
    assert data["counts"]["machine_only"] == 198  # 199 baseline - the 1 matched
    assert data["counts"]["duplicate_ccode_groups"] == 0

    matched = data["matched"][0]
    assert matched["ccode"] == "285278212"
    assert matched["baseline_name"] == "No. of blocks"
    assert matched["name_mismatch"] is False

    new = data["checklist_only"][0]
    assert new["ccode"] == "999999999"
    assert new["name"] == "Brand new param"
    # forward-filled from the previous row's Item column, same as a real
    # Excel with merged-looking category cells (only the first row of a
    # group actually has a value)
    assert new["category"] == "[Recipe]-[Group A]"


def test_api_secs_params_compare_excel_reports_duplicate_ccode(client):
    import base64

    rows = [
        ("[Recipe]-[Group A]", "First", 111111111, "DT_A"),
        (None, "First again, different name", 111111111, "DT_A2"),
    ]
    b64 = base64.b64encode(_checklist_xlsx_bytes(rows)).decode("ascii")
    res = client.post("/api/secs_params/compare_excel", json={"excel_base64": b64})
    assert res.status_code == 200
    data = res.get_json()
    assert data["counts"]["duplicate_ccode_groups"] == 1
    dup = data["duplicate_ccodes"][0]
    assert dup["ccode"] == "111111111"
    assert [r["name"] for r in dup["rows"]] == ["First", "First again, different name"]


def test_api_secs_params_compare_excel_requires_file(client):
    res = client.post("/api/secs_params/compare_excel", json={})
    assert res.status_code == 400


def test_api_secs_params_compare_excel_rejects_bad_base64(client):
    res = client.post("/api/secs_params/compare_excel", json={"excel_base64": "not valid base64!!"})
    assert res.status_code == 400


def test_api_secs_params_compare_excel_rejects_wrong_header(client):
    import base64
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Wrong", "Header", "Shape"])
    ws.append(["a", "b", "c"])
    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    res = client.post("/api/secs_params/compare_excel", json={"excel_base64": b64})
    assert res.status_code == 422
    assert "欄位標題" in res.get_json()["error"]


def test_api_strate_xml_download_zip_real_log(client):
    res = client.post("/api/strate_xml/download_zip", json={"log_base64": _secs_log_base64()})
    assert res.status_code == 200
    assert res.mimetype == "application/zip"
    assert "attachment" in res.headers["Content-Disposition"]

    import io
    import zipfile

    zf = zipfile.ZipFile(io.BytesIO(res.get_data()))
    names = zf.namelist()
    assert len(names) == 2
    assert all(n.endswith(".strate") for n in names)
    # Each entry must itself be a valid, parseable .strate file.
    from bingomap.strate import StrateFile

    for name in names:
        parsed = StrateFile.parse(zf.read(name).decode("utf-8"))
        assert parsed.die_info
